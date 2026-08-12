#!/usr/bin/env python3
"""build_workbook.py — assemble feature_catalog.xlsx from the registries."""
import csv
import os
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from nongrid_features import NONGRID

ROOT = Path(__file__).resolve().parent
SPEC = yaml.safe_load((ROOT / "feature_space.yaml").read_text())
# Output path is overridable so CI and local runs do not fight over a location.
OUT = os.environ.get("FEATURE_CATALOG_OUT", str(ROOT / "feature_catalog.xlsx"))
CANDIDATES = os.environ.get("FEATURE_CANDIDATES", str(ROOT / "candidates.csv"))

F = "Arial"
H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(name=F, bold=True, color="FFFFFF", size=10)
TITLE = Font(name=F, bold=True, size=14, color="1F3864")
SUB = Font(name=F, size=10, color="595959")
BODY = Font(name=F, size=10)
BOLD = Font(name=F, size=10, bold=True)
INPUT = Font(name=F, size=10, color="0000FF")
MONO = Font(name="Consolas", size=9)
THIN = Border(bottom=Side("thin", color="D9D9D9"))
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")

wb = Workbook()


def sheet(name, title, subtitle):
    ws = wb.create_sheet(name)
    ws["A1"] = title; ws["A1"].font = TITLE
    ws["A2"] = subtitle; ws["A2"].font = SUB
    ws.row_dimensions[2].height = 30
    ws["A2"].alignment = WRAP
    ws.sheet_view.showGridLines = False
    return ws


def header(ws, row, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=i, value=c)
        cell.font = H_FONT; cell.fill = H_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def write_rows(ws, start, rows, wrap_cols=()):
    for r, rec in enumerate(rows, start):
        for c, v in enumerate(rec, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = BODY; cell.border = THIN
            cell.alignment = WRAP if c in wrap_cols else TOP
    return start + len(rows)


# ══════════════════════════════════════════════════════════════════ README
ws = wb.active; ws.title = "README"
ws.sheet_view.showGridLines = False
widths(ws, {"A": 3, "B": 32, "C": 104})
ws["B2"] = "Chargeback Fraud — Feature Search Catalog"; ws["B2"].font = Font(name=F, bold=True, size=16, color="1F3864")
ws["B3"] = "Work queue for the feature-discovery agent. Generated from feature_space.yaml — regenerate, do not hand-edit structure."
ws["B3"].font = SUB

readme = [
 ("HOW THIS IS BUILT", ""),
 ("feature_space.yaml", "Source of truth. Slot registries: entities, measures, windows, filters, aggregations, transforms. Edit here."),
 ("expand_catalog.py", "Expands the YAML into candidates.csv. Default: --level L0, Stage A only. Re-run with --expand --survivors <file> after the univariate screen, and --level L1/L2 only where the Refinement Log justifies it."),
 ("nongrid_features.py", "Hand-enumerated families the grammar cannot express. Expected to carry most incremental lift."),
 ("build_workbook.py", "Regenerates this workbook."),
 ("", ""),
 ("SHEETS", ""),
 ("Grid Catalog", "Stage-A candidates: entity x measure x aggregation x window. One row per feature. Agent fills status / iv / psi_12m / coverage_pct / shap_rank / drop_reason."),
 ("Non-Grid Catalog", "Graph, novelty, deviation, consistency, strings, sequence, biometrics, instrument, basket, population."),
 ("Slot Registries", "The six slots, with priority and the attributes that drive gating."),
 ("Resolution Ladder", "L0/L1/L2 rungs, the nested-window correlation floor that justifies them, per-rung cost, and the four refinement triggers."),
 ("Refinement Log", "One row per marginal-gain test. Log negatives too — that is what stops the next person re-running them."),
 ("Typology Coverage", "Fraud-type x family matrix. Any zero cell is a coverage gap and must be closed or explicitly waived."),
 ("Data Sources", "Required inputs, whether we have them, and the integration backlog ranked by lift-per-effort."),
 ("Leakage Audit", "Mandatory checks. A feature that has not passed all of these may not ship."),
 ("Pruning Funnel", "Eight-stage screen with live counts. Record survivors at each stage."),
 ("", ""),
 ("STATUS VALUES", "candidate -> built -> screened -> shipped, or dropped (drop_reason required)"),
 ("CELL COLOURS", "Blue text = agent fills these in. Black = generated, do not overwrite."),
 ("", ""),
 ("NON-NEGOTIABLES", ""),
 ("Point-in-time", "Every label-dependent aggregate filters on label_arrival_ts, never event_ts. This is the #1 source of inflated offline AUC and it concentrates in the champion/challenger disagreement region."),
 ("Serving parity", "A feature that cannot be computed in the real-time path from the same code is a training-only feature and must be dropped or re-implemented as a streaming aggregate."),
 ("Censoring", "Entity outcome rates are computable only over approved traffic. Carry block_rate alongside every cb_rate and de-bias using release-program labels."),
 ("Shrinkage", "No raw rate on n below the row's min_support_n. Use empirical-Bayes toward the parent entity and expose n as a companion feature."),
 ("Operating point", "Final importance ranking is computed on the score band around the decision threshold, not globally."),
 ("Coverage vs resolution", "Entities, measures and non-grid families are COVERAGE — search exhaustively at every rung. Windows, transforms, filters and entity granularity are RESOLUTION — start at L0 and refine only on evidence."),
]
r = 5
for k, v in readme:
    ws.cell(row=r, column=2, value=k).font = BOLD if k and not v else (Font(name=F, bold=True, size=11, color="1F3864") if k and not v else BODY)
    if k and v:
        ws.cell(row=r, column=2).font = BOLD
    c = ws.cell(row=r, column=3, value=v); c.font = BODY; c.alignment = WRAP
    r += 1

# ══════════════════════════════════════════════════════ GRID CATALOG
if not os.path.exists(CANDIDATES):
    raise SystemExit(
        f"{CANDIDATES} not found.\n"
        "It is a build artifact, not source. Generate it first:\n"
        "    make catalog            # or: python expand_catalog.py --level L0\n"
        "Or build both in one step:  make workbook"
    )
rows = list(csv.reader(open(CANDIDATES)))
hdr, data = rows[0], rows[1:]
ws = sheet("Grid Catalog", "Stage-A grid candidates",
           "entity x measure x aggregation x window, filter=none, transform=raw. Screen these first; expand transforms and filters only "
           "onto survivors (expand_catalog.py --expand --survivors). Blue columns are for the agent to complete.")
header(ws, 4, hdr)
fill_cols = {hdr.index(c) + 1 for c in ("status", "drop_reason", "iv", "psi_12m", "coverage_pct", "shap_rank")}
for r_i, rec in enumerate(data, 5):
    for c_i, v in enumerate(rec, 1):
        cell = ws.cell(row=r_i, column=c_i, value=v)
        cell.font = INPUT if c_i in fill_cols else (MONO if c_i == 1 else BODY)
        cell.alignment = TOP
widths(ws, {"A": 46, "B": 6, "C": 10, "D": 30, "E": 15, "F": 13, "G": 26, "H": 15, "I": 20, "J": 14,
            "K": 12, "L": 16, "M": 14, "N": 30, "O": 44, "P": 15, "Q": 20, "R": 40, "S": 12, "T": 22,
            "U": 8, "V": 10, "W": 13, "X": 11})
ws.auto_filter.ref = f"A4:{get_column_letter(len(hdr))}{4 + len(data)}"
ws.cell(row=4, column=1).comment = Comment(
    "Naming contract: {entity}__{measure}__{agg}__{window}[__f_{filter}][__t_{transform}]", "catalog")

# ══════════════════════════════════════════════════ NON-GRID CATALOG
ws = sheet("Non-Grid Catalog", "Non-grid feature families",
           "Features the slot grammar cannot express. These carry most incremental lift and must be enumerated by hand. "
           "This list is the agent's starting point, not the finished set — extend it.")
ng_hdr = ["family", "feature_name", "definition", "inputs_required", "typologies_covered",
          "availability", "leakage_risk", "adversarial_halflife", "status", "drop_reason",
          "iv", "psi_12m", "coverage_pct", "shap_rank"]
header(ws, 4, ng_hdr)
for r_i, rec in enumerate(NONGRID, 5):
    for c_i, v in enumerate(list(rec) + ["candidate", "", "", "", "", ""], 1):
        cell = ws.cell(row=r_i, column=c_i, value=v)
        cell.font = INPUT if c_i >= 9 else (MONO if c_i == 2 else BODY)
        cell.alignment = WRAP if c_i in (3, 4) else TOP
widths(ws, {"A": 14, "B": 34, "C": 58, "D": 22, "E": 24, "F": 12, "G": 10, "H": 18,
            "I": 12, "J": 22, "K": 8, "L": 10, "M": 13, "N": 11})
ws.auto_filter.ref = f"A4:N{4 + len(NONGRID)}"

# ══════════════════════════════════════════════════ SLOT REGISTRIES
ws = sheet("Slot Registries", "Slot registries",
           "The six slots of the generative grammar. Priority drives tier gating in expand_catalog.py. "
           "adversarial_cost drives the half-life estimate: free-tier entities decay fastest and need the most frequent retraining.")
r = 4


def block(ws, r, title, cols, recs, wrapc=()):
    ws.cell(row=r, column=1, value=title).font = Font(name=F, bold=True, size=11, color="1F3864")
    r += 1
    header(ws, r, cols)
    ws.freeze_panes = None
    r += 1
    r = write_rows(ws, r, recs, wrapc)
    return r + 2


ents = SPEC["entities"]
r = block(ws, r, "ENTITIES — group-by keys", ["id", "name", "class", "adversarial_cost", "availability", "priority"],
          [(e["id"], e["name"], e["class"], e["adversarial_cost"], e["availability"], e.get("priority")) for e in ents])
r = block(ws, r, "COMPOSITE ENTITIES — each needs a stated fraud hypothesis",
          ["id", "parts", "hypothesis"],
          [(c["id"], " + ".join(c["parts"]), c["hypothesis"]) for c in SPEC["composites"]], wrapc=(3,))
r = block(ws, r, "MEASURES — what gets aggregated (label_dependent = point-in-time critical)",
          ["id", "name", "family", "label_dependent", "dtype", "priority"],
          [(m["id"], m["name"], m["family"], str(m["label_dependent"]), m["dtype"], m.get("priority")) for m in SPEC["measures"]])
r = block(ws, r, "WINDOWS — geometric ladder so adjacent ratios are informative",
          ["id", "label", "type", "seconds", "realtime", "priority"],
          [(w["id"], w["label"], w["type"], w["seconds"], str(w["realtime"]), w.get("priority")) for w in SPEC["windows"]])
r = block(ws, r, "FILTERS — same_* narrows; diff_* is often the signal itself",
          ["id", "label", "priority"], [(f["id"], f["label"], f.get("priority")) for f in SPEC["filters"]])
r = block(ws, r, "AGGREGATIONS", ["id", "name", "applies_to", "priority"],
          [(a["id"], a["name"], ", ".join(a["applies_to"]), a.get("priority")) for a in SPEC["aggregations"]])
r = block(ws, r, "TRANSFORMS — generalizes 'difference/ratio at different lags'",
          ["id", "name", "formula", "priority"],
          [(t["id"], t["name"], t["formula"], t.get("priority")) for t in SPEC["transforms"]], wrapc=(3,))
r = block(ws, r, "LABELS — model more than one; they have near-disjoint predictors",
          ["id", "definition", "maturity_days_90pct", "notes"],
          [(x["id"], x["definition"], x["maturity_days_90pct"], x["notes"]) for x in SPEC["labels"]], wrapc=(2, 4))
widths(ws, {"A": 22, "B": 40, "C": 46, "D": 18, "E": 22, "F": 10})

# ══════════════════════════════════════════════════ RESOLUTION LADDER
ws = sheet("Resolution Ladder", "Coarse-to-fine search control",
           "Coverage slots (entities, measures, non-grid families) are NEVER gated — a missing quantity is a blind spot no "
           "resolution recovers. Resolution slots start coarse and refine on evidence. Regenerate any rung with "
           "expand_catalog.py --level L0|L1|L2.")
r = 4
ws.cell(row=r, column=1, value="Why coarse-to-fine is safe on windows").font = Font(name=F, bold=True, size=11, color="1F3864")
r += 1
for line in [
 "For nested count windows under a homogeneous Poisson arrival process, Cov(N_short, N_long) = Var(N_short), so:",
 "        corr(N_short, N_long) = SQRT( w_short / w_long )",
 "Real transaction streams are burstier and autocorrelated, so measured correlation runs ABOVE this floor — treat it as a lower",
 "bound and measure your own. Windows spaced under ~x4 are >0.5 correlated and stage-6 redundancy clustering collapses them anyway,",
 "so fine resolution up front pays compute, storage and screening budget to build features you will then delete.",
 "",
 "Corollary 1  The ratio transform DECORRELATES. {raw_long, ratio_short/long} carries far more independent signal than two nested raws.",
 "Corollary 2  w30x7 (30d ending 7d ago) is the disjoint complement of w7d — uncorrelated by construction. One disjoint pair beats three nested.",
]:
    c = ws.cell(row=r, column=1, value=line)
    c.font = MONO if line.strip().startswith("corr") else BODY
    r += 1
r += 1

ws.cell(row=r, column=1, value="Nested-window correlation floor").font = Font(name=F, bold=True, size=11, color="1F3864")
r += 1
header(ws, r, ["spacing factor", "corr floor  =1/SQRT(k)", "verdict"]); r += 1
for k, v in [(2, "near-duplicate; stage 6 will delete one"), (3, "still heavily redundant"),
             (4, "minimum defensible spacing"), (8, "L1 target spacing"),
             (16, "L0 target spacing"), (30, "effectively independent")]:
    ws.cell(row=r, column=1, value=k).font = BODY
    ws.cell(row=r, column=2, value=f"=1/SQRT(A{r})").font = BODY
    ws.cell(row=r, column=2).number_format = "0.00"
    ws.cell(row=r, column=3, value=v).font = BODY
    r += 1
r += 2

LAD = SPEC["resolution_ladder"]
ws.cell(row=r, column=1, value="Rungs — windows by measure family").font = Font(name=F, bold=True, size=11, color="1F3864")
r += 1
header(ws, r, ["level", "outcome", "velocity", "behavior", "score", "transforms", "filters"]); r += 1
for lvl in ("L0", "L1", "L2"):
    cfg = LAD[lvl]
    w_ = cfg["windows"]
    fmt = lambda v: "ALL" if v == "ALL" else ", ".join(v)
    for c_i, v in enumerate([lvl, fmt(w_["outcome"]), fmt(w_["velocity"]), fmt(w_["behavior"]),
                             fmt(w_["score"]), fmt(cfg["transforms"]), fmt(cfg["filters"])], 1):
        cell = ws.cell(row=r, column=c_i, value=v)
        cell.font = BOLD if c_i == 1 else BODY
        cell.alignment = WRAP
    r += 1
r += 2

ws.cell(row=r, column=1, value="Cost per rung (tier-1, measured)").font = Font(name=F, bold=True, size=11, color="1F3864")
r += 1
header(ws, r, ["level", "A base", "+B transforms", "+C filters", "TOTAL", "note"]); r += 1
cost_start = r
for lvl, a_, b_, c_, note in [
    ("L0", 1167, 242, 0, "Probe. Run across the FULL coverage space. ~1 week of screening."),
    ("L1", 3016, 7128, 8734, "Standard. Only after L0 identifies the live quantities."),
    ("L2", 3918, 36470, 42689, "Full registry. Per-quantity refinement only — never bump globally.")]:
    ws.cell(row=r, column=1, value=lvl).font = BOLD
    for c_i, v in enumerate([a_, b_, c_], 2):
        ws.cell(row=r, column=c_i, value=v).font = BODY
    ws.cell(row=r, column=5, value=f"=SUM(B{r}:D{r})").font = BOLD
    ws.cell(row=r, column=6, value=note).font = BODY
    r += 1
ws.cell(row=r, column=1, value="ratio L2/L0").font = SUB
ws.cell(row=r, column=5, value=f"=E{cost_start+2}/E{cost_start}").font = SUB
ws.cell(row=r, column=5).number_format = "0.0x"
ws.cell(row=r, column=6, value="Windows are the SECOND-order saving — the dominant multiplier is B/C expansion, already deferred to post-screen.").font = SUB
r += 3

ws.cell(row=r, column=1, value="Refinement triggers — all must hold before adding resolution").font = Font(name=F, bold=True, size=11, color="1F3864")
r += 1
header(ws, r, ["#", "trigger", "rule"]); r += 1
trigs = SPEC["refinement_triggers"]
order = sorted(trigs, key=lambda t: 0 if t.get("priority") == "check_first" else 1)
for i, t in enumerate(order, 1):
    ws.cell(row=r, column=1, value=i).font = BODY
    ws.cell(row=r, column=2, value=t["id"] + ("  (check first)" if t.get("priority") else "")).font = BOLD
    c = ws.cell(row=r, column=3, value=" ".join(t["rule"].split())); c.font = BODY; c.alignment = WRAP
    ws.row_dimensions[r].height = 42
    r += 1
r += 1
ws.cell(row=r, column=1, value="Resolution is a PER-QUANTITY attribute. Card-testing velocity wants sub-minute; synthetic-identity age wants "
                               "multi-year. Never bump a whole family because one quantity needed it.").font = Font(name=F, size=10, italic=True, color="C00000")
widths(ws, {"A": 16, "B": 30, "C": 74, "D": 26, "E": 16, "F": 62, "G": 26})

# ══════════════════════════════════════════════════ REFINEMENT LOG
ws = sheet("Refinement Log", "Marginal-gain tests",
           "One row per refinement test. Log the result EITHER WAY — a recorded negative is what stops the next person re-running it. "
           "Pre-register the threshold before the test, not after.")
ws["A4"] = "Pre-registered thresholds"; ws["A4"].font = Font(name=F, bold=True, size=11, color="1F3864")
for i, (lab, val, note) in enumerate([
        ("min delta-AUC to keep", 0.0015, "on holdout, at the decision threshold band"),
        ("min delta-net $/month", 25000, "preferred criterion where the cost curve is trusted"),
        ("max added p99 latency (ms)", 2.0, "per refinement, against the 40ms vector budget")]):
    rr = 5 + i
    ws.cell(row=rr, column=1, value=lab).font = BODY
    c = ws.cell(row=rr, column=2, value=val); c.font = INPUT
    c.fill = PatternFill("solid", fgColor="FFFF00")
    ws.cell(row=rr, column=3, value=note).font = SUB

hdr_r = 10
header(ws, hdr_r, ["quantity (entity__measure__agg)", "from level", "tested addition", "edge_of_grid?",
                   "survived_screen?", "scale_gradient?", "delta_AUC", "delta_net_$/mo",
                   "added p99 ms", "verdict", "date", "notes"])
example = ["device__d_pan__dcnt", "L0", "add w1d between w1h and w30d", "no", "yes", "yes",
           0.0031, 41000, 0.8, "", "2026-08-11", "EXAMPLE ROW — delete before use. Gradient was steep 1h->30d, so the midpoint was tested."]
for c_i, v in enumerate(example, 1):
    cell = ws.cell(row=hdr_r + 1, column=c_i, value=v)
    cell.font = Font(name=F, size=10, italic=True, color="808080")
    cell.alignment = WRAP
ws.cell(row=hdr_r + 1, column=10,
        value=f'=IF(AND(G{hdr_r+1}>=$B$5,I{hdr_r+1}<=$B$7),"KEEP","DROP")').font = Font(name=F, size=10, italic=True, color="808080")
for rr in range(hdr_r + 2, hdr_r + 40):
    for c_i in range(1, 13):
        ws.cell(row=rr, column=c_i).font = INPUT
    ws.cell(row=rr, column=10, value=f'=IF(A{rr}="","",IF(AND(G{rr}>=$B$5,I{rr}<=$B$7),"KEEP","DROP"))').font = BODY
widths(ws, {"A": 36, "B": 11, "C": 34, "D": 14, "E": 17, "F": 16, "G": 11, "H": 15,
            "I": 13, "J": 10, "K": 12, "L": 54})
ws.auto_filter.ref = f"A{hdr_r}:L{hdr_r+40}"

# ══════════════════════════════════════════════════ TYPOLOGY COVERAGE
ws = sheet("Typology Coverage", "Fraud typology x feature family coverage",
           "Chargebacks are a mixed label. Each typology has near-disjoint predictors. Counts are live COUNTIFS over the "
           "Non-Grid Catalog. Any zero is a coverage gap: close it or waive it in writing.")
typs = SPEC["typologies"]
fams = sorted({r[0] for r in NONGRID})
header(ws, 4, ["typology"] + fams + ["TOTAL", "gap?"])
for i, t in enumerate(typs):
    rr = 5 + i
    ws.cell(row=rr, column=1, value=t["name"]).font = BOLD
    for j, fam in enumerate(fams):
        col = 2 + j
        ws.cell(row=rr, column=col,
                value=f'=COUNTIFS(\'Non-Grid Catalog\'!$A$5:$A${4+len(NONGRID)},"{fam}",'
                      f'\'Non-Grid Catalog\'!$E$5:$E${4+len(NONGRID)},"*{t["id"]}*")').font = BODY
    last = get_column_letter(1 + len(fams))
    ws.cell(row=rr, column=2 + len(fams), value=f"=SUM(B{rr}:{last}{rr})").font = BOLD
    ws.cell(row=rr, column=3 + len(fams),
            value=f'=IF(COUNTIF(B{rr}:{last}{rr},0)>0,"GAP: "&COUNTIF(B{rr}:{last}{rr},0)&" empty families","ok")').font = BODY
widths(ws, {"A": 26})
for j in range(len(fams)):
    ws.column_dimensions[get_column_letter(2 + j)].width = 13
ws.column_dimensions[get_column_letter(2 + len(fams))].width = 9
ws.column_dimensions[get_column_letter(3 + len(fams))].width = 30

note_r = 7 + len(typs)
ws.cell(row=note_r, column=1, value="Known structural gap — read before starting").font = Font(name=F, bold=True, size=11, color="C00000")
for i, line in enumerate([
    "First-party / friendly fraud shares almost no predictors with third-party fraud. If it is a material share of CB volume and the",
    "friendly-fraud row above is thin, that is the single largest gap in the feature set — not the absence of another velocity aggregate.",
    "Recommended: split the label by reason-code group and model the two populations separately (or as a multi-task head).",
    "Expect more lift from that split than from any additional entity aggregate."]):
    c = ws.cell(row=note_r + 1 + i, column=1, value=line); c.font = BODY

# ══════════════════════════════════════════════════ DATA SOURCES
ws = sheet("Data Sources", "Required inputs and integration backlog",
           "Ranked roughly by lift-per-integration-effort. Agent fills 'have_it' and 'owner' in the first pass — a feature whose "
           "input we do not have is a data request, not a dropped candidate.")
header(ws, 4, ["rank", "source", "unlocks", "families", "effort", "have_it?", "owner", "notes"])
sources = [
 (1, "Account event log (password/email/phone/payout changes, logins)", "The strongest ATO features. Most teams do not have these at score time.", "sequence, deviation", "med", "", "", "Needs streaming, not batch — ATO happens in minutes"),
 (2, "TC40 / SAFE issuer fraud feed", "Fast proxy label (7d vs 30d) AND an entity feature", "population, graph, labels", "med", "", "", "Also shortens the A/B readout cycle"),
 (3, "Device fingerprinting with cross-session linkage", "Device entity, graph edges, emulator/headless detection", "graph, biometrics, novelty", "med", "", "", "Linkage quality matters more than fingerprint entropy"),
 (4, "Entity resolution / graph store", "All graph features; component-level risk", "graph", "high", "", "", "Highest ceiling; ring fraud is invisible without it"),
 (5, "IP intelligence (datacenter/VPN/Tor, ASN abuse)", "Network-tier risk", "biometrics, consistency", "low", "", "", "Vendor; check latency budget"),
 (6, "Email intelligence (age in the wild, breach presence)", "Synthetic-identity detection", "novelty, strings", "low", "", "", "Vendor"),
 (7, "Phone intelligence (line type, carrier, port-in recency, tenure)", "Synthetic + ATO", "novelty, consistency", "low", "", "", "Port-in recency is an underused ATO signal"),
 (8, "Address normalization + freight-forwarder/reshipper lists", "Drop-address detection", "basket, consistency, graph", "low", "", "", "Cheap, high yield for physical goods"),
 (9, "Enriched BIN table (funding type, commercial, issuer, country)", "Instrument risk", "instrument", "low", "", "", "Ensure BIN-8 coverage, not just BIN-6"),
 (10, "3DS result data (attempt, auth status, liability shift)", "Separates friendly from third-party fraud", "instrument", "med", "", "", "Liability shift changes the economics, not just the risk"),
 (11, "Client telemetry SDK (keystroke, paste, autofill, focus)", "Behavioral biometrics", "biometrics, sequence", "med", "", "", "Paste-vs-type on the PAN field alone is worth the integration"),
 (12, "KYC document metadata (NFC vs photo, liveness, tamper score, doc reuse)", "Synthetic identity", "consistency, novelty, graph", "med", "", "", "Doc-number reuse across customers is a strong ring signal"),
 (13, "Dispute + customer-support contact history", "Friendly-fraud detection", "sequence, basket", "low", "", "", "Closes the biggest typology gap for most portfolios"),
 (14, "Logistics / delivery confirmation feed", "Friendly fraud (post-txn models only)", "basket", "med", "", "", "Not available at pre-auth; use for later-lifecycle scoring"),
 (15, "Consortium / network fraud data", "Cross-institution entity risk", "graph, population", "high", "", "", "Eligibility and contractual review required"),
 (16, "Marketing attribution (channel, campaign, affiliate)", "Fraud concentrates by affiliate", "sequence", "low", "", "", "Frequently available and almost never used"),
 (17, "Release-program labels (champion-blocked, sampled + released)", "De-biases every entity rate in the blocked region", "all outcome measures", "med", "", "", "Without this the model relearns the champion's blind spots"),
]
write_rows(ws, 5, sources, wrap_cols=(3, 8))
widths(ws, {"A": 6, "B": 52, "C": 50, "D": 26, "E": 8, "F": 10, "G": 14, "H": 44})
for rr in range(5, 5 + len(sources)):
    for cc in (6, 7):
        ws.cell(row=rr, column=cc).font = INPUT

# ══════════════════════════════════════════════════ LEAKAGE AUDIT
ws = sheet("Leakage Audit", "Mandatory correctness checks",
           "A feature that has not passed every applicable check may not ship, regardless of offline lift. "
           "Checks 1 and 2 are where nearly all inflated backtests come from.")
header(ws, 4, ["#", "check", "why it matters", "how to verify", "applies to", "pass?"])
checks = [
 (1, "Two-clock rule", "A CB on a 01 Jun txn reported 03 Jul must not appear in the 15 Jun entity rate. Hindsight labels inflate offline AUC and concentrate in the champion/challenger disagreement region — they flatter challengers specifically.", "Aggregates filter on label_arrival_ts, not event_ts. Recompute a sample of features as-of an historical timestamp and diff against the stored value.", "every label_dependent measure", ""),
 (2, "Serving parity", "Offline feature stores get mutated: labels arrive, records backfill, entities merge. The feature exists offline but its value contains hindsight.", "Diff shadow-computed values against offline replay on the same transactions. The diff IS the training/serving skew metric. Any feature above tolerance is rejected.", "all features", ""),
 (3, "Out-of-fold target encoding", "Entity rates computed on the same rows the model trains on are memorised, not learned.", "Time-forward or leave-one-out encoding scheme; verify by shuffling labels — encoded feature IV must collapse to ~0.", "all rate measures", ""),
 (4, "Censoring in the denominator", "A device the champion blocks 90% of the time looks clean, because only approvals produce CB labels. The feature encodes current policy and shifts the day thresholds change.", "Carry block_rate and decline_rate alongside every cb_rate. Re-estimate entity rates using release-program labels for the blocked region.", "all outcome measures", ""),
 (5, "Support floor + shrinkage", "A raw rate on n=1 is noise that trees will happily memorise.", "Enforce min_support_n from the catalog. Apply empirical-Bayes shrinkage toward the parent entity. Expose n as a companion feature.", "all rate measures", ""),
 (6, "Parent-entity hierarchy defined", "Shrinkage needs a target: BIN -> issuer -> country; IP -> /24 -> ASN; device -> device_model.", "Every rate feature names its parent in the catalog before it is built.", "all rate measures", ""),
 (7, "Latency budget", "Feature vector p99 must fit the decision budget. A 30-day warehouse scan is a training-only feature.", "Benchmark in the serving path, not offline. Re-implement as sketch-backed streaming aggregates (t-digest, count-min, HLL) or drop.", "all features", ""),
 (8, "Post-transaction contamination", "Delivery confirmation, refund outcome and dispute flags do not exist at pre-auth.", "Confirm each input's availability timestamp precedes the decision timestamp. Tag later-lifecycle-only features explicitly.", "basket, sequence", ""),
 (9, "Policy-encoding drift", "Features derived from champion score or block decisions shift when the operating point moves.", "Re-baseline these features after every threshold change; monitor their PSI separately.", "score family, block_rate", ""),
 (10, "Adversarial half-life tagged", "Free-tier features (email, UA, IP) decay in weeks. A model leaning on them backtests well and degrades fast.", "Every feature carries a half-life. Track per-feature lift decay and PSI as a standing monitor; set retraining cadence from the mix.", "all features", ""),
 (11, "PII and regulatory review", "Some fields are protected or prohibited as model inputs in some jurisdictions, and proxies for them can be too.", "Legal/compliance sign-off on birth country, national ID, name-derived and geo-derived features before they enter a production model.", "identity, geo classes", ""),
 (12, "Disparate-impact screen", "Geography- and name-derived features can proxy protected attributes.", "Run fairness diagnostics on the candidate set at the intended operating point, not just globally.", "identity, geo classes", ""),
]
write_rows(ws, 5, checks, wrap_cols=(3, 4, 5))
widths(ws, {"A": 5, "B": 26, "C": 62, "D": 62, "E": 26, "F": 8})
for rr in range(5, 5 + len(checks)):
    ws.cell(row=rr, column=6).font = INPUT

# ══════════════════════════════════════════════════ PRUNING FUNNEL
ws = sheet("Pruning Funnel", "Eight-stage screen",
           "Cheapest filter first. Record survivor counts; the drop reason for every eliminated feature goes in the catalog sheets. "
           "Counts in column D are live formulas over the two catalog sheets.")
header(ws, 4, ["stage", "screen", "criterion", "in", "out", "survivors", "notes"])
ngN = 4 + len(NONGRID); gridN = 4 + len(data)
stages = [
 ("0", "Enumerate", "Grid stage-A + non-grid families",
  f"=COUNTA('Grid Catalog'!$A$5:$A${gridN})+COUNTA('Non-Grid Catalog'!$B$5:$B${ngN})", "", "", "Starting population"),
 ("1", "Availability", "Computable at decision time within the latency budget", "", "", "", "Drops batch_only and vendor calls over budget"),
 ("2", "Support", "Median n per entity-window cell >= min_support_n", "", "", "", "Kills sparse composites automatically"),
 ("3", "Stability", "12-month PSI and coverage; no seasonal collapse or coverage cliff", "", "", "", "Fill psi_12m and coverage_pct in the catalogs"),
 ("4", "Univariate screen", "Information value / mutual information vs each label", "", "", "", "Cheap and parallel. Run per label, not just the primary"),
 ("5", "Transform + filter expansion", "Re-run expand_catalog.py --expand --survivors on stage-4 survivors", "", "", "", "Expanding before this point multiplies screening cost ~10x for no coverage gain"),
 ("6", "Redundancy", "Cluster by rank correlation; keep one per cluster", "", "", "", "Prefer the cheaper and more adversarially durable member of each cluster"),
 ("7", "Multivariate", "Permutation importance / SHAP on holdout, iterative tail drop", "", "", "", "Fill shap_rank"),
 ("8", "Leakage audit", "Every applicable check on the Leakage Audit sheet passes", "", "", "", "Anything suspiciously top-ranked gets a manual timeline inspection"),
 ("9", "Operating-point relevance", "Re-rank importance restricted to the score band around the threshold", "", "", "", "Global importance is dominated by the easy tail; the decision happens at the cut"),
]
write_rows(ws, 5, stages, wrap_cols=(3, 7))
widths(ws, {"A": 7, "B": 30, "C": 58, "D": 12, "E": 10, "F": 12, "G": 60})
for rr in range(5, 5 + len(stages)):
    for cc in (4, 5, 6):
        if ws.cell(row=rr, column=cc).value in (None, ""):
            ws.cell(row=rr, column=cc).font = INPUT

r = 7 + len(stages)
ws.cell(row=r, column=1, value="Definition of done").font = Font(name=F, bold=True, size=11, color="1F3864")
for i, line in enumerate([
 "1. Every candidate row has status != 'candidate'; every dropped row has a drop_reason.",
 "2. Typology Coverage has no GAP flag, or each gap carries a written waiver.",
 "3. Every shipped feature has passed all applicable Leakage Audit checks and has a named parent entity for shrinkage.",
 "4. Shadow-vs-offline value diff is measured and within tolerance for the full shipped vector.",
 "5. Data Sources sheet has have_it and owner filled for every row; missing inputs are open data requests with owners.",
 "6. The shipped set is re-ranked at the decision threshold, not globally, and the ranking is recorded.",
]):
    ws.cell(row=r + 1 + i, column=1, value=line).font = BODY

for name in wb.sheetnames:
    wb[name].sheet_properties.tabColor = "1F3864"
wb.save(OUT)
print("wrote", OUT)
print("grid rows:", len(data), "| non-grid rows:", len(NONGRID))
